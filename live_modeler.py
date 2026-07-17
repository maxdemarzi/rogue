from ontology import con
import openpyxl

def generate_live_excel(ticker="AIR", year=2012, filename="financial_model_live.xlsx", db_path="rogue_finance.duckdb"):
    """
    Generates a live Excel spreadsheet containing fundamental snapshots and 
    associated analyst calculations using actual Excel formulas.
    """
    # Query financial statements snapshot
    query = """
    SELECT
      ticker,
      fiscal_year,
      revenue,
      cogs,
      grossprofit,
      operating_income_loss,
      ebitda,
      interest_expense,
      total_debt,
      assets,
      liabilities,
      working_capital,
      retained_earnings,
      market_capitalization,
      accounts_receivable
    FROM fundamentals_snapshots
    WHERE ticker = ? AND fiscal_year = ?;
    """
    row = con.execute(query, [ticker, year]).fetchone()

    
    if not row:
        raise ValueError(f"No fundamentals snapshot found for {ticker} in year {year}.")
        
    (
        _, _, revenue, cogs, _, operating_income, ebitda, interest_expense,
        total_debt, assets, liabilities, working_capital, retained_earnings,
        market_cap, accounts_receivable
    ) = row
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{ticker} Modeler ({year})"
    
    # Enable gridlines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Set headers
    ws["A1"] = "Financial Modeler"
    ws["A1"].font = openpyxl.styles.Font(size=14, bold=True)
    
    # Modeler inputs (hardcoded values from DB)
    inputs = [
        ("Ticker", ticker),
        ("Fiscal Year", year),
        ("Revenue ($)", revenue),
        ("COGS ($)", cogs),
        ("Operating Income ($)", operating_income),
        ("EBITDA ($)", ebitda),
        ("Interest Expense ($)", interest_expense),
        ("Total Debt ($)", total_debt),
        ("Total Assets ($)", assets),
        ("Total Liabilities ($)", liabilities),
        ("Working Capital ($)", working_capital),
        ("Retained Earnings ($)", retained_earnings),
        ("Market Capitalization ($)", market_cap),
        ("Accounts Receivable ($)", accounts_receivable)
    ]
    
    # Write inputs starting from row 3
    for idx, (label, val) in enumerate(inputs):
        row_num = 3 + idx
        ws[f"A{row_num}"] = label
        ws[f"B{row_num}"] = val
        ws[f"A{row_num}"].font = openpyxl.styles.Font(bold=True)
        
    # Formulas (dynamically evaluated by Excel using actual Excel formula strings!)
    # Row indices for reference:
    # B3: Ticker, B4: Fiscal Year, B5: Revenue, B6: COGS, B7: Operating Income
    # B8: EBITDA, B9: Interest Expense, B10: Total Debt, B11: Total Assets, B12: Total Liabilities
    # B13: Working Capital, B14: Retained Earnings, B15: Market Capitalization, B16: Accounts Receivable
    
    formulas = [
        ("Gross Profit ($)", "=B5-B6"),
        ("Gross Profit Margin (%)", "=E3/B5"),
        ("EBITDA Margin (%)", "=B8/B5"),
        ("Debt-to-Equity (%)", "=B10/B14"),  # Total Debt / Market Capitalization (Equity proxy)
        ("Interest Coverage (x)", "=B7/B9"),
        ("Days Sales Outstanding (DSO)", "=365*B16/B5"),
        ("Altman Z-Score", "=1.2*(B13/B11) + 1.4*(B14/B11) + 3.3*(B7/B11) + 0.6*(B15/B12) + 0.999*(B5/B11)")
    ]
    
    ws["D2"] = "Live Modeler Formulas"
    ws["D2"].font = openpyxl.styles.Font(size=12, bold=True)
    
    for idx, (label, formula) in enumerate(formulas):
        row_num = 3 + idx
        ws[f"D{row_num}"] = label
        ws[f"E{row_num}"] = formula
        ws[f"D{row_num}"].font = openpyxl.styles.Font(bold=True)
        ws[f"E{row_num}"].font = openpyxl.styles.Font(italic=True)
        
    # Autofit column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(filename)
    print(f"Generated live Excel model: {filename}")
    return filename

if __name__ == "__main__":
    try:
        generate_live_excel()
    except Exception as e:
        print("Excel generation failed:", e)
